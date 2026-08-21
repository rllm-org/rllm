"""The GDPval weighted-rubric reward_fn.

Covers the three things that make it *this* grader rather than the cookbook's:
it reads the Stirrup harness's artifact keys, it carries the structural signals
the per-task verifier would otherwise have reported, and it refuses to score
what it cannot see (no rubric, no deliverable, audio/video).
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from rllm.eval.reward_fns import gdpval_rubric as rubric
from rllm.types import Episode, Task

RUBRIC = json.dumps(
    [
        {"score": 2, "criterion": "The deliverable is an Excel workbook.", "rubric_item_id": "a"},
        {"score": 3, "criterion": "The sample size is 220.", "rubric_item_id": "b"},
        {"score": -2, "criterion": "The deliverable fabricates figures.", "rubric_item_id": "c"},
    ]
)


def _task(**metadata) -> Task:
    base = {"prompt": "Produce a sampling plan.", "rubric_json": RUBRIC, "occupation": "Accountants", "sector": "Services"}
    return Task(id="task-1", instruction="Produce a sampling plan.", metadata={**base, **metadata})


def _manifest(tmp_path, termination_type, *, artifacts=1, rejected=()):
    path = tmp_path / "submission_manifest.json"
    path.write_text(
        json.dumps(
            {
                "termination": {"type": termination_type},
                "artifacts": [{"local_path": f"f{i}.xlsx"} for i in range(artifacts)],
                "rejected_paths": list(rejected),
            }
        )
    )
    return path


# --------------------------------------------------------------------------- #
# Rubric parsing / aggregation semantics
# --------------------------------------------------------------------------- #


def test_positive_weights_earn_and_negative_weights_penalize():
    criteria = rubric.parse_rubric(RUBRIC)
    assert [c.score for c in criteria] == [2, 3, -2]

    # Only the positive criteria count toward the denominator.
    assert rubric.aggregate(criteria, [True, True, False]).reward == pytest.approx(1.0)
    assert rubric.aggregate(criteria, [True, False, False]).reward == pytest.approx(2 / 5)

    # A penalty criterion firing subtracts from what was earned.
    assert rubric.aggregate(criteria, [True, True, True]).reward == pytest.approx(3 / 5)


def test_penalties_cannot_drive_the_reward_below_zero():
    """A run that earns nothing and trips every penalty still scores 0, not negative."""
    criteria = rubric.parse_rubric(RUBRIC)
    assert rubric.aggregate(criteria, [False, False, True]).reward == 0.0


# --------------------------------------------------------------------------- #
# Deliverable discovery against the Stirrup harness's artifact keys
# --------------------------------------------------------------------------- #


def test_reads_the_stirrup_harness_deliverables_key():
    """``deliverables`` is what rllm.harnesses.stirrup publishes."""
    episode = Episode(artifacts={"deliverables": ["/corpus/a.xlsx", "/corpus/b.docx"]})
    assert rubric._deliverable_candidate_paths(episode) == ["/corpus/a.xlsx", "/corpus/b.docx"]


def test_still_reads_the_older_cookbook_artifact_keys():
    """An episode from either harness must grade the same."""
    episode = Episode(artifacts={"deliverable_path": "/corpus/a.xlsx", "output_files": ["/corpus/b.docx"]})
    assert rubric._deliverable_candidate_paths(episode) == ["/corpus/a.xlsx", "/corpus/b.docx"]


def test_candidate_paths_are_deduplicated_across_keys():
    episode = Episode(artifacts={"deliverables": ["/corpus/a.xlsx"], "deliverable_path": "/corpus/a.xlsx"})
    assert rubric._deliverable_candidate_paths(episode) == ["/corpus/a.xlsx"]


# --------------------------------------------------------------------------- #
# Structural signals: a catalog reward_fn replaces the per-task verifier, so
# this grader has to report them or they vanish.
# --------------------------------------------------------------------------- #


def test_structural_signals_come_from_the_harness_manifest(tmp_path):
    episode = Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish", artifacts=2))})
    signals = {s.name: s.value for s in rubric._structural_signals(episode)}
    assert signals == {"finish_called": 1.0, "abandoned": 0.0, "submission_valid": 1.0, "artifact_count": 2.0}


def test_abandonment_is_distinguishable_from_a_bad_deliverable(tmp_path):
    """Both score 0; conflating them would hide *why*."""
    episode = Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "abandon_task_finish", artifacts=0))})
    signals = {s.name: s.value for s in rubric._structural_signals(episode)}
    assert signals["abandoned"] == 1.0
    assert signals["finish_called"] == 0.0
    assert signals["submission_valid"] == 0.0


def test_rejected_paths_make_a_submission_invalid(tmp_path):
    episode = Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish", rejected=["/etc/passwd"]))})
    signals = {s.name: s.value for s in rubric._structural_signals(episode)}
    assert signals["finish_called"] == 1.0
    assert signals["submission_valid"] == 0.0


def test_structural_signals_survive_a_missing_manifest():
    """A crashed run must still produce finite signals, never an exception."""
    signals = {s.name: s.value for s in rubric._structural_signals(Episode(artifacts={}))}
    assert signals == {"finish_called": 0.0, "abandoned": 0.0, "submission_valid": 0.0, "artifact_count": 0.0}


# --------------------------------------------------------------------------- #
# Fail-closed: never score what the judge cannot see
# --------------------------------------------------------------------------- #


def test_a_task_without_a_rubric_is_ungraded_not_zero(tmp_path):
    out = rubric.evaluate_rubric(_task(rubric_json=""), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata["reason"] == "no_rubric"
    assert out.metadata["ungraded"] is True
    # The structural facts still come through, so the run is not a black hole.
    assert {s.name for s in out.signals} >= {"finish_called", "artifact_count"}


def test_no_surfaced_deliverable_is_ungraded(tmp_path):
    """Fail-closed: never grade the transcript in place of the artifact."""
    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish", artifacts=0))}))
    assert out.metadata["reason"] == "no_deliverable_surfaced"
    assert out.reward == 0.0


def test_media_without_a_gemini_key_is_ungraded(tmp_path, monkeypatch):
    """Media needs Google's Files API; without a key, refuse rather than guess."""
    monkeypatch.delenv("GEMINI_API_KEY", raising=False)
    video = tmp_path / "walkthrough.mp4"
    video.write_bytes(b"\x00\x00\x00 ftypmp42")
    episode = Episode(artifacts={"deliverables": [str(video)], "submission_manifest": str(_manifest(tmp_path, "finish"))})

    out = rubric.evaluate_rubric(_task(), episode)
    assert out.metadata["reason"] == "media_requires_gemini_key"
    assert out.metadata["ungraded"] is True
    assert "walkthrough.mp4" in out.metadata["media_files"]
    assert {s.name: s.value for s in out.signals}["ungraded_media"] == 1.0


def test_is_media_only_matches_audio_and_video():
    assert rubric.is_media("a.mp4") and rubric.is_media("a.WAV")
    assert not rubric.is_media("a.xlsx")
    assert not rubric.is_media("a.pdf")


def test_no_judge_configured_is_ungraded(tmp_path, monkeypatch):
    """A missing judge is a harness problem, and must not read as a model failure."""
    xlsx = tmp_path / "Sample.xlsx"
    xlsx.write_text("not really a workbook")
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: (None, None, None))
    monkeypatch.setattr(rubric, "_find_deliverable_view", lambda task, ep: rubric.DeliverableView("some text", [], "file:Sample.xlsx"))

    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata["reason"] == "no_judge_configured"
    assert out.metadata["ungraded"] is True


# --------------------------------------------------------------------------- #
# Grading a deliverable end to end, with a stubbed judge
# --------------------------------------------------------------------------- #


def test_grades_a_deliverable_and_writes_an_audit_trail(tmp_path, monkeypatch):
    corpus = tmp_path / "corpus"
    corpus.mkdir()
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: ("stub-judge", None, None))
    monkeypatch.setattr(rubric, "_find_deliverable_view", lambda task, ep: rubric.DeliverableView("workbook text", [], "file:Sample.xlsx"))
    monkeypatch.setattr(rubric, "_find_input_view", lambda task: rubric.DeliverableView("", [], "none"))
    # First two criteria met, the penalty criterion not triggered.
    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", lambda *a, **k: ([True, True, False], ["ok", "ok", "absent"]))
    monkeypatch.setattr(rubric, "_verify_rubric", lambda criteria, met, just, judge: met)

    episode = Episode(artifacts={"submission_dir": str(corpus), "submission_manifest": str(_manifest(tmp_path, "finish", artifacts=2))})
    out = rubric.evaluate_rubric(_task(), episode)

    assert out.reward == pytest.approx(1.0)
    assert out.is_correct is True
    signals = {s.name: s.value for s in out.signals}
    assert signals["rubric_score"] == pytest.approx(1.0)
    assert signals["finish_called"] == 1.0  # structural signals ride along
    assert out.metadata["criteria_met"] == 2
    assert out.metadata["criteria_total"] == 3

    decisions = json.loads((corpus / "judge_decisions.json").read_text())
    assert decisions["task_id"] == "task-1"
    assert decisions["judge_model"] == "stub-judge"
    assert [c["met"] for c in decisions["criteria"]] == [True, True, False]


def test_judge_decisions_land_in_the_run_keyed_corpus_dir(tmp_path):
    """Not next to an arbitrary file: the corpus dir is what the arena will read."""
    corpus = tmp_path / "run" / "task__0"
    corpus.mkdir(parents=True)
    rubric._write_judge_decisions(Episode(artifacts={"submission_dir": str(corpus)}), {"reward": 0.5})
    assert json.loads((corpus / "judge_decisions.json").read_text())["reward"] == 0.5


# --------------------------------------------------------------------------- #
# Wiring
# --------------------------------------------------------------------------- #


def test_the_registry_name_resolves_to_this_grader():
    from rllm.eval.evaluator_loader import load_evaluator

    # Only the reward_fn name is registered. The dataset's public name is
    # deliberately not an alias: ``dataset.toml`` writes this exact string, so
    # a second spelling would be one nothing resolves through.
    assert load_evaluator("gdpval_rubric_reward_fn") is not None


def test_gdpval_has_one_public_rubric_graded_catalog_entry():
    import importlib.resources

    catalog = json.loads((importlib.resources.files("rllm") / "registry" / "datasets.json").read_text())["datasets"]
    assert catalog["gdpval"]["reward_fn"] == "gdpval_rubric_reward_fn"
    assert "gdpval-rubric" not in catalog


def test_this_module_exports_no_solver_system_prompt():
    """A verifier hint would corrupt the byte-exact AA solver prompt."""
    assert not hasattr(rubric, "SYSTEM_PROMPT")


# --------------------------------------------------------------------------- #
# Unreadable deliverables: a broken host must not read as a bad model
# --------------------------------------------------------------------------- #


def test_extraction_failure_is_detected():
    """Missing parsers leave only an error marker where content should be."""
    text = "### File: Sample.xlsx\n[extraction error for Sample.xlsx: No module named 'openpyxl']"
    assert rubric._extraction_failed(text)


def test_partial_extraction_is_still_gradable():
    """One unreadable attachment among readable ones leaves real content to judge."""
    text = "### File: a.xlsx\nSheet1: revenue 220\n### File: b.zip\n[extraction error for b.zip: bad zip]"
    assert not rubric._extraction_failed(text)


def test_readable_text_is_not_flagged():
    assert not rubric._extraction_failed("### File: a.xlsx\nSheet1: revenue 220")
    assert not rubric._extraction_failed("")


def test_an_unreadable_deliverable_is_ungraded_not_scored_near_zero(tmp_path, monkeypatch):
    """The failure that matters: handed an error string, a judge answers "No" to
    every content criterion and the run scores a plausible ~0.03 that looks like
    a bad deliverable instead of a broken grader."""
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: ("stub-judge", None, None))
    monkeypatch.setattr(
        rubric,
        "_find_deliverable_view",
        lambda task, ep: rubric.DeliverableView(
            "### File: Sample.xlsx\n[extraction error for Sample.xlsx: No module named 'openpyxl']",
            [],
            "file:Sample.xlsx",
        ),
    )

    def _must_not_be_called(*a, **k):
        raise AssertionError("the judge must not be called on an unreadable deliverable")

    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", _must_not_be_called)

    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata["reason"] == "deliverable_unreadable"
    assert out.metadata["ungraded"] is True
    assert "gdpval" in out.metadata["hint"]


def test_rendered_page_images_keep_an_unparsed_deliverable_gradable(tmp_path, monkeypatch):
    """Text extraction can fail while the page images still show the document."""
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: ("stub-judge", None, None))
    monkeypatch.setattr(
        rubric,
        "_find_deliverable_view",
        lambda task, ep: rubric.DeliverableView("[extraction error for a.pdf: boom]", ["data:image/jpeg;base64,AAA"], "file:a.pdf"),
    )
    monkeypatch.setattr(rubric, "_find_input_view", lambda task: rubric.DeliverableView("", [], "none"))
    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", lambda *a, **k: ([True, False, False], ["ok", "no", "no"]))
    monkeypatch.setattr(rubric, "_verify_rubric", lambda criteria, met, just, judge: met)

    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata.get("reason") is None
    assert out.reward == pytest.approx(2 / 5)


def test_the_renderer_uses_the_published_gdpval_image():
    """Every evaluator uses the same pinned LibreOffice and font closure."""
    from rllm.data import gdpval_aa

    assert rubric._render_docker_image() == gdpval_aa.published_image_ref()


def test_office_conversion_runs_only_in_the_pinned_container(tmp_path, monkeypatch):
    import subprocess

    from rllm.data.gdpval_aa import AA_PLATFORM

    source = tmp_path / "report.xlsx"
    source.write_bytes(b"workbook")
    calls = []

    def fake_run(command, **kwargs):
        calls.append((command, kwargs))
        output_mount = next(arg for arg in command if arg.startswith("type=bind,source=") and arg.endswith(",target=/output"))
        output_dir = output_mount.removeprefix("type=bind,source=").removesuffix(",target=/output")
        (Path(output_dir) / "report.pdf").write_bytes(b"pdf")
        return subprocess.CompletedProcess(command, 0)

    monkeypatch.setattr(subprocess, "run", fake_run)

    assert rubric._soffice_convert(source, "pdf") == b"pdf"
    assert len(calls) == 1
    command, kwargs = calls[0]
    assert command[:7] == ["docker", "run", "--rm", "--platform", AA_PLATFORM, "--network", "none"]
    assert rubric._render_docker_image() in command
    assert f"type=bind,source={tmp_path},target=/input,readonly" in command
    assert command[-1] == "/input/report.xlsx"
    assert kwargs == {"check": True, "capture_output": True, "timeout": 180}


# --------------------------------------------------------------------------- #
# Audio/video: routed to Gemini's Files API, which can actually watch them
# --------------------------------------------------------------------------- #


def test_a_media_deliverable_is_graded_through_the_gemini_judge(tmp_path, monkeypatch):
    """With a key, the chunk/retry/verify machinery runs unchanged on media."""
    from rllm.eval.reward_fns import _gdpval_media as media

    monkeypatch.setenv("GEMINI_API_KEY", "test-key")
    video = tmp_path / "walkthrough.mp4"
    video.write_bytes(b"\x00\x00\x00 ftypmp42")

    uploaded = []
    monkeypatch.setattr(media, "upload_media", lambda p, key, **k: uploaded.append(p) or {"file_uri": "files/abc", "mime_type": "video/mp4", "name": "files/abc"})
    monkeypatch.setattr(rubric, "upload_media", media.upload_media)
    monkeypatch.setattr(rubric, "ffprobe_metadata", lambda p: "# Media file: walkthrough.mp4\n- duration_s: 42")
    monkeypatch.setattr(rubric, "make_media_judge", lambda model, key, files, **k: lambda *a, **kw: '[{"rubric_id": 1, "status": "Yes", "justification": "seen"}]')
    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", lambda *a, **k: ([True, True, False], ["seen", "seen", "no"]))
    monkeypatch.setattr(rubric, "_verify_rubric", lambda criteria, met, just, judge: met)

    episode = Episode(artifacts={"deliverables": [str(video)], "submission_manifest": str(_manifest(tmp_path, "finish"))})
    out = rubric.evaluate_rubric(_task(), episode)

    assert out.metadata.get("reason") is None
    assert out.reward == pytest.approx(1.0)
    assert out.metadata["judge_model"] == media.media_judge_model()
    assert out.metadata["media_files"] == ["walkthrough.mp4"]


def test_a_failed_media_upload_is_ungraded_not_scored(tmp_path, monkeypatch):
    """A file the judge never received must not be scored from its metadata."""
    from rllm.eval.reward_fns._gdpval_media import MediaUploadError

    monkeypatch.setenv("GEMINI_API_KEY", "test-key")
    video = tmp_path / "walkthrough.mp4"
    video.write_bytes(b"\x00\x00\x00 ftypmp42")

    def _boom(*a, **k):
        raise MediaUploadError("ingestion ended in state FAILED")

    monkeypatch.setattr(rubric, "upload_media", _boom)
    monkeypatch.setattr(rubric, "ffprobe_metadata", lambda p: "# Media file: walkthrough.mp4")

    episode = Episode(artifacts={"deliverables": [str(video)], "submission_manifest": str(_manifest(tmp_path, "finish"))})
    out = rubric.evaluate_rubric(_task(), episode)
    assert out.metadata["reason"] == "media_upload_failed"
    assert out.metadata["ungraded"] is True


def test_oversize_media_is_refused_before_upload(tmp_path, monkeypatch):
    from rllm.eval.reward_fns import _gdpval_media as media

    monkeypatch.setenv("GDPVAL_MEDIA_MAX_MB", "1")
    big = tmp_path / "big.mp4"
    big.write_bytes(b"x" * (2 * 1024 * 1024))
    with pytest.raises(media.MediaUploadError, match="over the"):
        media.upload_media(big, "test-key")


def test_uploaded_media_becomes_a_file_data_part(tmp_path):
    """The judge references the uploaded file, not inline bytes."""
    from rllm.eval.reward_fns._gdpval_media import _to_gemini_contents

    messages = [
        {"role": "system", "content": "you are a judge"},
        {"role": "user", "content": [{"type": "text", "text": "grade this"}, {"type": "image_url", "image_url": {"url": "data:image/jpeg;base64,AAA"}}]},
    ]
    system, contents = _to_gemini_contents(messages, [{"file_uri": "files/abc", "mime_type": "video/mp4"}])

    assert system == {"parts": [{"text": "you are a judge"}]}
    parts = contents[0]["parts"]
    assert {"file_data": {"file_uri": "files/abc", "mime_type": "video/mp4"}} in parts
    assert {"inline_data": {"mime_type": "image/jpeg", "data": "AAA"}} in parts


def test_media_extension_detection():
    from rllm.eval.reward_fns._gdpval_media import is_media, media_mime

    assert is_media("a.mp4") and is_media("a.WAV") and is_media("a.opus")
    assert not is_media("a.xlsx")
    assert media_mime("a.mp4").startswith("video/")


def test_ffprobe_metadata_is_readable_when_the_probe_fails(monkeypatch):
    """Metadata is best-effort; an unavailable probe must not raise."""
    from rllm.eval.reward_fns import _gdpval_media as media

    monkeypatch.setattr(media, "_ffprobe_json", lambda p: None)
    assert "metadata unavailable" in media.ffprobe_metadata("a.mp4")


def test_the_render_image_carries_libreoffice_and_ffmpeg():
    """One image serves both office->PDF and ffprobe, from AA's own closure."""
    from rllm.data import gdpval_aa
    from rllm.eval.reward_fns._gdpval_media import render_docker_image

    assert render_docker_image() == gdpval_aa.published_image_ref()


def test_a_non_vision_judge_is_reported_as_text_only(tmp_path, monkeypatch):
    """A downgrade must not be reported as multimodal grading.

    Fireworks' glm-5p2 rejects image inputs, so the judge silently falls back to
    text. Reporting the *prepared* image count then claims the page renderings
    were graded when every visual criterion was decided from text alone.
    """
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: ("stub-judge", None, None))
    monkeypatch.setattr(rubric, "_find_deliverable_view", lambda task, ep: rubric.DeliverableView("workbook text", ["data:image/jpeg;base64,AAA"] * 3, "file:Sample.xlsx"))
    monkeypatch.setattr(rubric, "_find_input_view", lambda task: rubric.DeliverableView("", [], "none"))
    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", lambda *a, **k: ([True, False, False], ["ok", "no", "no"]))
    monkeypatch.setattr(rubric, "_verify_rubric", lambda criteria, met, just, judge: met)

    def _fell_back_judge(model, base_url, api_key):
        def call(messages, text_only_messages=None):
            return "[]"

        call.text_only_fallbacks = 1  # a multimodal call was rejected
        return call

    monkeypatch.setattr(rubric, "_make_rubric_judge", _fell_back_judge)

    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata["deliverable_images_prepared"] == 3
    assert out.metadata["deliverable_images"] == 0
    assert out.metadata["vision_used"] is False
    assert out.metadata["judge_text_only_fallback"] is True


def test_a_vision_judge_reports_the_images_it_consumed(tmp_path, monkeypatch):
    monkeypatch.setattr(rubric, "_resolve_judge", lambda task: ("stub-judge", None, None))
    monkeypatch.setattr(rubric, "_find_deliverable_view", lambda task, ep: rubric.DeliverableView("workbook text", ["data:image/jpeg;base64,AAA"] * 3, "file:Sample.xlsx"))
    monkeypatch.setattr(rubric, "_find_input_view", lambda task: rubric.DeliverableView("", [], "none"))
    monkeypatch.setattr(rubric, "_grade_rubric_multimodal", lambda *a, **k: ([True, False, False], ["ok", "no", "no"]))
    monkeypatch.setattr(rubric, "_verify_rubric", lambda criteria, met, just, judge: met)

    out = rubric.evaluate_rubric(_task(), Episode(artifacts={"submission_manifest": str(_manifest(tmp_path, "finish"))}))
    assert out.metadata["deliverable_images"] == 3
    assert out.metadata["vision_used"] is True
    assert out.metadata["judge_text_only_fallback"] is False


# --------------------------------------------------------------------------- #
# Image ordering and the single Gemini request
# --------------------------------------------------------------------------- #


def test_images_are_consecutive_by_file_and_page():
    picked = rubric._flatten_images([["a0", "a1"], ["b0", "b1"], ["c0"]], cap=10)
    assert picked == ["a0", "a1", "b0", "b1", "c0"]


def test_flattening_handles_no_images():
    assert rubric._flatten_images([], cap=10) == []
    assert rubric._flatten_images([[], []], cap=10) == []


def test_default_image_limit_is_500(monkeypatch):
    monkeypatch.delenv("GDPVAL_JUDGE_MAX_IMAGES", raising=False)
    assert rubric._max_judge_images() == 500


def test_truncation_is_logged_not_silent(caplog):
    """A dropped page must leave a trace; silent caps read as full coverage."""
    with caplog.at_level("WARNING"):
        assert rubric._flatten_images([["a0", "a1", "a2"]], cap=2) == ["a0", "a1"]
    assert any("page images available" in record.getMessage() for record in caplog.records)


def test_input_files_are_extracted_and_rendered_in_file_order(tmp_path, monkeypatch):
    task_dir = tmp_path / "task-1" / "environment" / "files"
    task_dir.mkdir(parents=True)
    first = task_dir / "first.xlsx"
    second = task_dir / "second.pdf"
    first.write_bytes(b"first")
    second.write_bytes(b"second")
    task = _task(reference_files=[first.name, second.name])
    task.dataset_dir = tmp_path
    task.sub_dir = Path("task-1")

    monkeypatch.setattr(rubric, "extract_deliverable_text", lambda path, max_chars=None: f"### File: {Path(path).name}\ntext")
    monkeypatch.setattr(
        rubric,
        "render_deliverable_images",
        lambda path: [f"{Path(path).stem}-page-1", f"{Path(path).stem}-page-2"],
    )

    view = rubric._find_input_view(task)

    assert view.images == ["first-page-1", "first-page-2", "second-page-1", "second-page-2"]
    assert view.text.index("### File: first.xlsx") < view.text.index("### File: second.pdf")
    assert "first.xlsx is rendered as 2 consecutive page image(s)" in view.text
    assert "second.pdf is rendered as 2 consecutive page image(s)" in view.text


def test_input_files_resolve_from_in_sandbox_absolute_paths(tmp_path):
    """``reference_files`` holds sandbox paths; grading runs on the host.

    The builder writes ``/home/user/<name>`` because the solver prompt must
    quote the paths the solver will see. Joining those onto the staged dir
    discards the base (pathlib resets on an absolute right operand) and hands
    back a path that never exists on the host — which silently drops every
    input file from the judge's context instead of failing.
    """
    staged = tmp_path / "task-1" / "environment" / "files"
    staged.mkdir(parents=True)
    (staged / "outline.docx").write_bytes(b"outline")

    task = _task(reference_files=["/home/user/outline.docx"])
    task.dataset_dir = tmp_path
    task.sub_dir = Path("task-1")

    resolved = rubric._input_file_paths(task)

    assert resolved == [str(staged / "outline.docx")]
    assert Path(resolved[0]).exists()


def test_input_and_output_images_share_one_cap(monkeypatch):
    monkeypatch.setenv("GDPVAL_JUDGE_MAX_IMAGES", "3")
    criteria = rubric.parse_rubric(RUBRIC)
    seen = []

    def judge(mm, txt=None):
        seen.extend(part["image_url"]["url"] for part in mm[1]["content"] if part["type"] == "image_url")
        return json.dumps(
            [
                {"rubric_id": 1, "status": "No"},
                {"rubric_id": 2, "status": "No"},
                {"rubric_id": 3, "status": "No"},
            ]
        )

    rubric._grade_rubric_multimodal(
        criteria,
        rubric.DeliverableView("output", ["output-1", "output-2"]),
        "prompt",
        judge,
        "input",
        ["input-1", "input-2"],
    )

    assert seen == ["input-1", "input-2", "output-1"]


def test_all_images_are_sent_in_one_call():
    criteria = rubric.parse_rubric(RUBRIC)
    seen: list[list[str]] = []

    def judge(mm, txt=None):
        images = [part["image_url"]["url"] for part in mm[1]["content"] if part.get("type") == "image_url"]
        seen.append(images)
        return json.dumps(
            [
                {"rubric_id": 1, "status": "No", "justification": "x"},
                {"rubric_id": 2, "status": "Yes", "justification": "seen"},
                {"rubric_id": 3, "status": "No", "justification": "x"},
            ]
        )

    results = rubric._grade_one_segment(criteria, "text", ["p0", "p1", "p2", "p3"], "prompt", judge)

    assert seen == [["p0", "p1", "p2", "p3"]]
    assert results[2]["hit"] == 1
    assert results[2]["justification"] == "seen"


def test_prompt_sections_and_attachments_are_in_the_discussed_order():
    criteria = rubric.parse_rubric(RUBRIC)
    mm, text_only = rubric._grading_messages(
        "do the task",
        "### File: result.pdf\ncontent",
        ["page-1", "page-2"],
        list(enumerate(criteria, 1)),
        "### File: source.csv\nvalue",
        ["input-page-1", "input-page-2"],
    )

    text_parts = [part["text"] for part in mm[1]["content"] if part["type"] == "text"]
    assert text_parts[0].startswith("## 1. Task Prompt")
    assert text_parts[1].startswith("## 2. Input Files")
    assert "2 input page image(s) are attached immediately below" in text_parts[1]
    assert text_parts[2].lstrip().startswith("---\n\n## 3. Model Output")
    assert "2 page image(s) are attached immediately below" in text_parts[2]
    assert text_parts[3].lstrip().startswith("---\n\n## 4. Rubric Items")
    content_types = [part["type"] for part in mm[1]["content"]]
    assert content_types == ["text", "text", "image_url", "image_url", "text", "image_url", "image_url", "text"]
    assert [part["image_url"]["url"] for part in mm[1]["content"] if part["type"] == "image_url"] == [
        "input-page-1",
        "input-page-2",
        "page-1",
        "page-2",
    ]
    body = text_only[1]["content"]
    assert body.index("## 1. Task Prompt") < body.index("## 2. Input Files") < body.index("## 3. Model Output") < body.index("## 4. Rubric Items")


# --------------------------------------------------------------------------- #
# Spreadsheet recalculation
# --------------------------------------------------------------------------- #


def _workbook(path, b3):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"] = "Population", 200
    ws["A2"], ws["B2"] = "Buffer", 20
    ws["A3"], ws["B3"] = "Sample size", b3
    wb.save(path)
    return path


def test_workbook_markdown_nests_file_then_sheet(tmp_path):
    text = rubric.extract_deliverable_text(_workbook(tmp_path / "sample.xlsx", 220), max_chars=None)
    assert text.startswith("### File: sample.xlsx\n#### Sheet: Sheet1\n")


def test_uncached_formulas_are_detected(tmp_path):
    """openpyxl writes the formula but no computed result."""
    assert rubric._has_uncached_formulas(_workbook(tmp_path / "f.xlsx", "=B1+B2")) is True


def test_a_workbook_of_literals_needs_no_recalculation(tmp_path):
    assert rubric._has_uncached_formulas(_workbook(tmp_path / "v.xlsx", 220)) is False


def test_recalculation_recovers_a_formula_value(tmp_path, monkeypatch):
    """Without this the cell reads blank and every numeric criterion fails."""
    source = _workbook(tmp_path / "src.xlsx", "=B1+B2")
    recalculated = _workbook(tmp_path / "recalc.xlsx", 220)
    monkeypatch.setattr(rubric, "_soffice_convert", lambda p, ext: recalculated.read_bytes())

    text = rubric._extract_xlsx(source)
    assert "220" in text


def test_a_failed_conversion_keeps_the_original_read(tmp_path, monkeypatch):
    """No converter available must degrade, not raise."""
    source = _workbook(tmp_path / "src.xlsx", "=B1+B2")
    monkeypatch.setattr(rubric, "_soffice_convert", lambda p, ext: None)

    text = rubric._extract_xlsx(source)
    assert "Population" in text  # still readable, just missing the computed cell


def test_a_conversion_that_loses_content_does_not_win(tmp_path, monkeypatch):
    """Prefer the recalculated read only when it recovered cells."""
    source = _workbook(tmp_path / "src.xlsx", "=B1+B2")
    from openpyxl import Workbook

    empty = tmp_path / "empty.xlsx"
    Workbook().save(empty)
    monkeypatch.setattr(rubric, "_soffice_convert", lambda p, ext: empty.read_bytes())

    text = rubric._extract_xlsx(source)
    assert "Population" in text
